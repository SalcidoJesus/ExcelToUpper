import sys
from openpyxl import load_workbook
from os.path import isfile


print("""
---------------------------------------
-                                     -
-   Conversor a mayúsculas en Excel   -
-                                     -
---------------------------------------
""")


#nombre archivo
try:
    nombreArchivo = input("-Nombre del archivo a convertir (incluir extención):\n")
    if nombreArchivo == "":
        exit()
    elif not isfile(nombreArchivo):
        print("Archivo no existente")
        exit()
except:
    print("Nombre no válido")
    exit()

wb = load_workbook(filename= nombreArchivo)

#hoja
try:
    nombreHoja = input("-Nombre de la hoja a utilizar:\n")
    if nombreHoja == "":
        exit()
except:
    print("Hoja no válida")
    exit()

#buscar hoja
try:
    hoja = wb[nombreHoja]
except:
    print("hoja no encontrada")
    exit()

#nombre archivo nuevo
try:
    nombreArchivoNuevo = input("-Nombre del archivo nuevo: (incluir extención)\n")
    if nombreArchivoNuevo == "":
        exit()
except:
    print("Nombre no válido")
    exit()



#columna mínina
try:
    minCol = int(input("-Columna mínima a leer:\n"))
    if minCol < 1:
        print("La columna mímina será 1")
        minCol = 1
except:
    print("Ingresaste un número no válido")
    exit()

#columna máxima
try:
    maxCol = int(input("-Columna máxima a leer:\n"))
    if maxCol < minCol:
        print("La columna máxima no puede ser menor a la mínima")
        sys.exit(0)
except:
    print("ingresaste un número no válido")
    exit()

#fila mínina
try:
    minRow = int(input("-Fila mínima a leer:\n"))
    if minRow == "" or minRow < 1:
        print("La columna mímina será 1")
        minRow = 1
except:
    print("Ingresaste un número no válido")
    exit()

#fila máxima
try:
    maxRow = int(input("-Fila máxima a leer:\n"))
    if maxRow < minRow:
        print("La fila máxima debe ser igual o mayor a la fila mínima")
        exit()
except:
    print("ingresaste un número no válido")
    exit()
"""
# print("minCol:", minCol)
# print("maxCol:", maxCol)
# print("minRow:", minRow)
# print("maxRow:", maxRow)
"""
for iRow in range(minRow, maxRow + 1):
    # print("--- Leyendo datos de la fila {}".format(iRow))
    for iCol in range(minCol, maxCol + 1):
        # print("celda {}-{}".format(iRow, iCol))
        celda = hoja.cell(row = iRow, column = iCol)
        
        if isinstance(celda.value, str):
            celda.value = celda.value.upper()

        # if celda.value is None:
        #     print("*vacio")
        # else:
        #     print(celda.value)

try:
    wb.save(filename=nombreArchivoNuevo)
    print("""
    ---------------------------------------
    -    Archivo guardado con éxito :D    -
    ---------------------------------------
    """)
except:
    print("algo salió mal")
